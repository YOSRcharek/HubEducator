from django.db.models import Count, Q
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from .forms import AddUserForm, EditUserForm, EditUserForm
from core.models import CertificateAttempt, User
from django.contrib import messages
from django.core.paginator import Paginator
from django.contrib import messages
from core.forms import ProfileUpdateForm
from django.contrib.auth import get_user_model
from django.shortcuts import render, get_object_or_404, redirect
from django.views import View
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from core.models import Speciality
from .forms import SpecialityForm
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from core.models import CertificatExercise, Certificate
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from core.models import Certificate, CertificatExercise, Speciality
from .forms import CertificateForm, CertificatExerciseForm
from django.forms import modelformset_factory
from django.db import transaction
import csv
from django.http import HttpResponse
from openpyxl import Workbook
@login_required
def dashboard(request):
    if request.user.role not in ['admin']:
        return redirect(reverse('unauthorized'))  # 'unauthorized' is the URL name
    return render(request, 'dashboard.html', {})

@login_required
def users(request):
    if request.user.role not in ['admin']:
        return redirect(reverse('unauthorized'))
    
    users_list = User.objects.exclude(role='admin')  # exclure admin si tu veux
    paginator = Paginator(users_list, 3)  # 5 utilisateurs par page

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'users/users.html', {'page_obj': page_obj})

@login_required
def adduser(request):
    if request.user.role not in ['admin', 'teacher']:
        return redirect(reverse('unauthorized'))

    if request.method == "POST":
        form = AddUserForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('users')  # après ajout, retour à la liste des users
    else:
        form = AddUserForm()

    return render(request, 'users/adduser.html', {'form': form})

@login_required
def edit_user(request, user_id):
    if request.user.role != "admin":
        return redirect(reverse("unauthorized"))

    user = get_object_or_404(User, pk=user_id)

    if request.method == "POST":
        form = EditUserForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            return redirect("users")  # retourne vers la liste des users
    else:
        form = EditUserForm(instance=user)

    return render(request, "users/edituser.html", {"form": form, "user_obj": user})

@login_required
def delete_user(request, user_id):
    if request.user.role not in ['admin']:
        return redirect(reverse('unauthorized'))
    user = get_object_or_404(User, id=user_id)
    user.delete()
    messages.success(request, "User deleted successfully!")
    return redirect('users')  


@method_decorator(login_required, name='dispatch')
class SpecialityListView(View):
    def get(self, request):
        specialities = Speciality.objects.all()
        return render(request, 'specialities/listSpecialities.html', {'specialities': specialities})

@method_decorator(login_required, name='dispatch')
class SpecialityCreateView(View):
    def get(self, request):
        form = SpecialityForm()
        return render(request, 'specialities/addSpecialities.html', {'form': form})

    def post(self, request):
        form = SpecialityForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('specialities')  # corriger nom url
        return render(request, 'specialities/addSpecialities.html', {'form': form})

@method_decorator(login_required, name='dispatch')
class SpecialityUpdateView(View):
    def get(self, request, speciality_id):
        speciality = get_object_or_404(Speciality, pk=speciality_id)
        form = SpecialityForm(instance=speciality)
        return render(request, 'specialities/editSpecialities.html', {'form': form})

    def post(self, request, speciality_id):
        speciality = get_object_or_404(Speciality, pk=speciality_id)
        form = SpecialityForm(request.POST, instance=speciality)
        if form.is_valid():
            form.save()
            return redirect('specialities')
        return render(request, 'specialities/editSpecialities.html', {'form': form})

@method_decorator(login_required, name='dispatch')
class SpecialityDeleteView(View):
    def post(self, request, speciality_id):
        speciality = get_object_or_404(Speciality, pk=speciality_id)
        speciality.delete()
        return redirect('specialities')

class ListCertificatView(View):
    def get(self, request):
        certificates_qs = Certificate.objects.annotate(
            exercise_count=Count('exercises'),
            participant_count=Count('certificateattempt', distinct=True),
            succeeded_count=Count('certificateattempt', filter=Q(certificateattempt__passed=True), distinct=True)
        ).all().order_by('id')
        paginator = Paginator(certificates_qs, 10)  # 10 items per page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        return render(request, 'certificats/listCertificate.html', {
            'certificates': page_obj.object_list,
            'page_obj': page_obj
        })


def create_certificate(request):
    # Formset pour les exercices, permettant la suppression
    CertificatExerciseFormSet = modelformset_factory(CertificatExercise, form=CertificatExerciseForm, extra=1, can_delete=True)

    if request.method == 'POST':
        cert_form = CertificateForm(request.POST, request.FILES)
        # Ne PAS passer request.FILES au formset (pas de champs fichiers dans CertificatExercise)
        formset = CertificatExerciseFormSet(request.POST, queryset=CertificatExercise.objects.none(), prefix='exercise')

        if cert_form.is_valid() and formset.is_valid():
            certificate = cert_form.save()
            for exercise_form in formset:
                if exercise_form.cleaned_data.get('DELETE'):
                    if exercise_form.instance.pk:
                        exercise_form.instance.delete()
                else:
                    exercise = exercise_form.save(commit=False)
                    exercise.certificate = certificate
                    exercise.save()
            return redirect('list_certificates')
        else:
            # Debug: afficher erreurs si la validation échoue pour aider au diagnostic
            from django.contrib import messages
            messages.error(request, f"Certificate form errors: {cert_form.errors}")
            messages.error(request, f"Formset errors: {formset.errors}")
            print("Certificate form errors:", cert_form.errors)
            print("Formset errors:", formset.errors)
    else:
        cert_form = CertificateForm()
        formset = CertificatExerciseFormSet(queryset=CertificatExercise.objects.none(), prefix='exercise')

    return render(request, 'certificats/certificate_form.html', {
        'cert_form': cert_form,
        'formset': formset,
    })

# AJOUT / MODIF: edit_certificate (passer et traiter formset avec exercises existants)
@login_required
def edit_certificate(request, cert_id):
    certificate = get_object_or_404(Certificate, pk=cert_id)
    
    CertificatExerciseFormSet = modelformset_factory(
        CertificatExercise,
        form=CertificatExerciseForm,
        extra=0,  # pas de form vide par défaut pour l'édition
        can_delete=True
    )

    if request.method == "POST":
        cert_form = CertificateForm(request.POST, request.FILES, instance=certificate)
        # Ne PAS passer request.FILES au formset
        formset = CertificatExerciseFormSet(
            request.POST,
            queryset=CertificatExercise.objects.filter(certificate=certificate),
            prefix='exercise'
        )
        
        if cert_form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    certificate = cert_form.save()
                    for form in formset.forms:
                        if form.cleaned_data.get('DELETE'):
                            if form.instance.pk:
                                form.instance.delete()
                        else:
                            exercise = form.save(commit=False)
                            exercise.certificate = certificate
                            exercise.save()
                    
                    messages.success(request, "Certificate updated successfully!")
                    return redirect('list_certificates')
            except Exception as e:
                messages.error(request, f"Error updating certificate: {str(e)}")
                print(f"Error: {str(e)}")
        else:
            # Si invalid, afficher erreurs pour debug
            messages.error(request, f"Certificate form errors: {cert_form.errors}")
            messages.error(request, f"Formset errors: {formset.errors}")
            print("Edit certificate form errors:", cert_form.errors)
            print("Edit formset errors:", formset.errors)
    else:
        cert_form = CertificateForm(instance=certificate)
        formset = CertificatExerciseFormSet(
            queryset=CertificatExercise.objects.filter(certificate=certificate),
            prefix='exercise'
        )

    context = {
        'cert_form': cert_form,
        'formset': formset,
        'certificate': certificate
    }
    return render(request, 'certificats/editCerificate.html', context)

# AJOUT: delete_certificate
@login_required
def delete_certificate(request, cert_id):
    if request.user.role not in ['admin', 'teacher']:
        return redirect(reverse('unauthorized'))
    certificate = get_object_or_404(Certificate, pk=cert_id)
    if request.method == "POST":
        certificate.delete()
        messages.success(request, "Certificate deleted successfully.")
        return redirect('list_certificates')
    # If not POST, redirect back to list
    return redirect('list_certificates')

@login_required
def preview_certificate(request, cert_id):
    if request.user.role not in ['admin', 'teacher']:
        return redirect(reverse('unauthorized'))
    certificate = get_object_or_404(Certificate, pk=cert_id)
    exercises = CertificatExercise.objects.filter(certificate=certificate)
    return render(request, 'certificats/preview_certificate.html', {
        'certificate': certificate,
        'exercises': exercises,
    })

@login_required
def certificate_results(request, cert_id):
    if request.user.role not in ['admin', 'teacher']:
        return redirect(reverse('unauthorized'))
    certificate = get_object_or_404(Certificate, pk=cert_id)
    attempts = CertificateAttempt.objects.filter(certificate=certificate).select_related('user').order_by('-completed_at')
    return render(request, 'certificats/certificate_results.html', {
        'certificate': certificate,
        'attempts': attempts,
    })

@login_required
def attempt_details(request, attempt_id):
    if request.user.role not in ['admin', 'teacher']:
        return redirect(reverse('unauthorized'))
    attempt = get_object_or_404(CertificateAttempt, pk=attempt_id)
    answers = attempt.answers.select_related('exercise').all()
    return render(request, 'certificats/attempt_details.html', {
        'attempt': attempt,
        'answers': answers,
    })

@login_required
def export_certificate_results_csv(request, cert_id):
    if request.user.role not in ['admin', 'teacher']:
        return redirect(reverse('unauthorized'))
    certificate = get_object_or_404(Certificate, pk=cert_id)
    attempts = CertificateAttempt.objects.filter(certificate=certificate).select_related('user').prefetch_related('answers__exercise').order_by('-completed_at')

    # Get all exercises for this certificate
    exercises = list(CertificatExercise.objects.filter(certificate=certificate).order_by('id'))

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Certificate Results"

    # Write headers
    headers = ['User', 'Score', 'Total Questions', 'Passed', 'Completed At']
    for exercise in exercises:
        headers.append(f"{exercise.question} - Answer")
        headers.append(f"{exercise.question} - Correct Answer")

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data
    for row_num, attempt in enumerate(attempts, 2):
        ws.cell(row=row_num, column=1, value=attempt.user.username)
        ws.cell(row=row_num, column=2, value=attempt.score)
        ws.cell(row=row_num, column=3, value=attempt.total_questions)
        ws.cell(row=row_num, column=4, value='Yes' if attempt.passed else 'No')
        ws.cell(row=row_num, column=5, value=attempt.completed_at.strftime('%Y-%m-%d %H:%M:%S'))

        # Create a dict of exercise_id to answer for quick lookup
        answers_dict = {answer.exercise_id: answer for answer in attempt.answers.all()}

        col_num = 6
        for exercise in exercises:
            if exercise.id in answers_dict:
                answer = answers_dict[exercise.id]
                ws.cell(row=row_num, column=col_num, value=answer.answer)
                ws.cell(row=row_num, column=col_num + 1, value=exercise.correct_answer)
            else:
                ws.cell(row=row_num, column=col_num, value='')
                ws.cell(row=row_num, column=col_num + 1, value=exercise.correct_answer)
            col_num += 2

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{certificate.title}_results.xlsx"'

    wb.save(response)
    return response
