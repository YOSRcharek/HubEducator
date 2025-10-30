from django.db.models import Count, Q
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from django.contrib import messages
from django.core.paginator import Paginator
from urllib3 import request
from .forms import AddUserForm, EditUserForm
from core.models import User, Course,CourseCategory, Lesson, SubLesson, Resource,Review
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.utils import timezone
from django.db.models import Avg
from core.models import User, Course, CertificateAttempt, Speciality, CertificatExercise, Certificate
from django.views import View
from django.utils.decorators import method_decorator
from .forms import SpecialityForm, CertificateForm, CertificatExerciseForm
from django.forms import modelformset_factory
from django.db import transaction
import csv
from django.http import HttpResponse
from openpyxl import Workbook
from django.views.decorators.http import require_http_methods
from core.models import User, Course,CourseCategory, Lesson, SubLesson, Resource
from etude.models import GroupeEtude, ResourceEtude
from etude.forms import GroupeEtudeForm
from etude.forms import ResourceEtudeForm
from etude.models import Meeting
from .forms import MeetingForm
from etude.google_utils import get_google_credentials_for_user

import uuid
from django.http import JsonResponse
import datetime
from django.utils.dateparse import parse_date, parse_time
import uuid
from django.utils import timezone as dj_timezone


# --------------------------
# Teacher Dashboard
# --------------------------
@login_required
def TeacherDash(request):
    if request.user.role != 'teacher':
        return redirect(reverse('unauthorized'))
    return render(request, 'teacherDash.html', {})


# --------------------------
# List Students assigned to teacher's courses
# --------------------------
@login_required
def students(request):
    if request.user.role != 'teacher':
        return redirect(reverse('unauthorized'))

    # Récupérer tous les étudiants inscrits aux cours de cet enseignant
    students_list = User.objects.filter(
        role='student',
        enrolled_courses__teacher=request.user
    ).distinct()

    paginator = Paginator(students_list, 5)  # 5 étudiants par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'students/students.html', {'page_obj': page_obj})


# --------------------------
# Add a new student
# --------------------------
@login_required
def add_student(request):
    if request.user.role not in ['admin', 'teacher']:
        return redirect(reverse('unauthorized'))

    if request.method == "POST":
        form = AddUserForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save(commit=False)
            user.role = 'student'  # s'assurer que c'est un student
            user.set_password(form.cleaned_data['password1'])
            user.save()
            messages.success(request, "Student added successfully!")
            return redirect('students')  # redirige vers la liste des students
    else:
        form = AddUserForm()

    return render(request, 'students/addstudent.html', {'form': form})


# --------------------------
# Edit a student
# --------------------------
@login_required
def edit_student(request, user_id):
    if request.user.role not in ['admin', 'teacher']:
        return redirect(reverse("unauthorized"))

    student = get_object_or_404(User, pk=user_id)

    if request.method == "POST":
        form = EditUserForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, "Student updated successfully!")
            return redirect("students")
    else:
        form = EditUserForm(instance=student)

    return render(request, "students/editstudent.html", {"form": form, "student_obj": student})


# --------------------------
# Delete a student
# --------------------------
@login_required
def delete_student(request, user_id):
    if request.user.role not in ['admin', 'teacher']:
        return redirect(reverse('unauthorized'))

    student = get_object_or_404(User, id=user_id)
    student.delete()
    messages.success(request, "Student deleted successfully!")
    return redirect('students')

@login_required
def student_detail(request, user_id):
    if request.user.role != 'teacher':
        return redirect(reverse('unauthorized'))

    student = get_object_or_404(User, pk=user_id)

    # Récupérer les cours de l’étudiant liés à cet enseignant
    courses = student.enrolled_courses.filter(teacher=request.user)

    return render(request, 'students/student_detail.html', {
        'student': student,
        'courses': courses
    })



@login_required
def courses(request):
    if request.user.role != 'teacher':
        return redirect('unauthorized')

    # Tous les cours du prof
    teacher_courses = Course.objects.filter(teacher=request.user)

    # Mettre à jour le status des cours
    for course in teacher_courses:
        if course.lessons.exists() and course.status == 'pending':
            course.status = 'inprogress'
            course.save()
        elif not course.lessons.exists() and course.status != 'pending':
            course.status = 'pending'
            course.save()
        elif course.max_lessons and course.lessons.count() >= course.max_lessons:
            course.status = 'completed'
            course.save()

    # Filtre par status
    status_filter = request.GET.get('status', '')
    if status_filter:
        courses_list = teacher_courses.filter(status=status_filter)
    else:
        courses_list = teacher_courses

    # Pagination
    paginator = Paginator(courses_list.distinct(), 6)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Stats
    current_courses_count = teacher_courses.count()
    active_courses = teacher_courses.filter(status='inprogress').count()
    completed_courses = teacher_courses.filter(status='completed').count()
    pending_courses = teacher_courses.filter(status='pending').count()

    # 🔹 Tous les étudiants assignés aux cours du prof
    my_students = User.objects.filter(
        enrolled_courses__teacher=request.user,
        role='student'
    ).distinct()

    my_students_count = my_students.count()

    return render(request, 'courses/courses.html', {
        'page_obj': page_obj,
        'status_options': ['pending', 'inprogress', 'completed'],
        'current_status': status_filter,
        'current_courses_count': current_courses_count,
        'active_courses': active_courses,
        'completed_courses': completed_courses,
        'pending_courses': pending_courses,
        'my_students': my_students,
        'my_students_count': my_students_count,
    })

@login_required
def add_courses(request):
    categories = CourseCategory.objects.all()

    if request.method == 'POST':
        errors = []

        title = request.POST.get('title', '').strip()
        if not title:
            errors.append("Course title is required.")

        # Capacity
        capacity = request.POST.get('capacity')
        try:
            capacity = int(capacity)
            if capacity <= 0:
                errors.append("Capacity must be a positive number.")
        except:
            capacity = 30

        # Category
        category_name = request.POST.get('category')
        category_obj = CourseCategory.objects.filter(name=category_name).first() if category_name else None
        if not category_obj:
            errors.append("Please select a valid category.")

        # Dates
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if not start_date:
            errors.append("Start date is required.")
        if not end_date:
            errors.append("End date is required.")
        if start_date and end_date and start_date > end_date:
            errors.append("End date must be after start date.")

        # Thumbnail
        thumbnail_file = request.FILES.get('thumbnail')
        if not thumbnail_file:
            errors.append("Course thumbnail is required.")

        # si erreurs -> réafficher form avec messages
        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'courses/addCourses.html', {'categories': categories})

        # ✅ Pas d'erreur, créer le cours
        description = request.POST.get('description', '').strip()
        max_lessons = request.POST.get('max_lessons')
        course = Course.objects.create(
            title=title,
            description=description,
            capacity=capacity,
            category=category_obj,
            teacher=request.user,
            status='pending',
            start_date=start_date,
            end_date=end_date,
            max_lessons=int(max_lessons)
        )


        course.thumbnail.save(thumbnail_file.name, thumbnail_file)

        # Lessons (optionnel)
        max_sublessons_list = request.POST.getlist('max_sublessons[]')  # <-- récupère tous les max sublessons      
        lesson_titles = request.POST.getlist('lesson_title[]')
        lesson_descriptions = request.POST.getlist('lesson_description[]')
        lessons_dict = {}

        for idx, title_lesson in enumerate(lesson_titles, start=1):
            if not title_lesson.strip():
                continue  # skip empty lesson titles
            desc = lesson_descriptions[idx-1] if idx-1 < len(lesson_descriptions) else ''

            try:
                max_sub = int(max_sublessons_list[idx-1]) if idx-1 < len(max_sublessons_list) and max_sublessons_list[idx-1] else None
            except:
                max_sub = None

            lesson_obj = Lesson.objects.create(
                course=course,
                title=title_lesson,
                description=desc,
                order=idx,
                max_sublessons=max_sub  # <-- ici
            )
            lessons_dict[idx] = lesson_obj

            # Fichiers uploadés
            for file in request.FILES.getlist(f'lesson_resources_{idx}[]'):
                Resource.objects.create(lesson=lesson_obj, sub_lesson=None, title=file.name, resource_type='pdf', file=file)

            # Ressources externes
            external_urls = request.POST.getlist('external_url[]')
            resource_types = request.POST.getlist('resource_type[]')
            for url, rtype in zip(external_urls, resource_types):
                if url.strip():
                    Resource.objects.create(
                        lesson=lesson_obj,
                        sub_lesson=None,
                        title=url.split("/")[-1][:30],
                        resource_type=rtype,
                        external_url=url
                    )


        # SubLessons (optionnel)
        sublesson_titles = request.POST.getlist('sublesson_title[]')
        sublesson_contents = request.POST.getlist('sublesson_content[]')
        lesson_objs = list(lessons_dict.values())
        for idx, sub_title in enumerate(sublesson_titles, start=1):
            if not sub_title.strip():
                continue
            content = sublesson_contents[idx-1] if idx-1 < len(sublesson_contents) else ''
            attach_lesson = lesson_objs[(idx-1) % len(lesson_objs)] if lesson_objs else None
            sublesson_obj = SubLesson.objects.create(lesson=attach_lesson, title=sub_title, content=content)

            # Fichiers uploadés
            for file in request.FILES.getlist(f'sublesson_resources_{idx}[]'):
                Resource.objects.create(lesson=None, sub_lesson=sublesson_obj, title=file.name, resource_type='pdf', file=file)

            # Ressources externes
            external_urls = request.POST.getlist('external_url[]')
            resource_types = request.POST.getlist('resource_type[]')
            for url, rtype in zip(external_urls, resource_types):
                if url.strip():
                    Resource.objects.create(
                        lesson=None,
                        sub_lesson=sublesson_obj,
                        title=url.split("/")[-1][:30],
                        external_url=url,     
                        resource_type='external'  
                    )

        messages.success(request, "Course created successfully!")
        return redirect('courses')

    return render(request, 'courses/addCourses.html', {'categories': categories})

@login_required
def course_edit(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    categories = CourseCategory.objects.all()

    if request.method == 'POST':
        # Update basic fields
        course.title = request.POST.get('title', '').strip()
        course.description = request.POST.get('description', '').strip()

        capacity = request.POST.get('capacity')
        if capacity and capacity.isdigit():
            course.capacity = int(capacity)

        course.level = request.POST.get('level') or course.level
        course.start_date = request.POST.get('start_date') or course.start_date
        course.end_date = request.POST.get('end_date') or course.end_date

        category_name = request.POST.get('category')
        category_obj = CourseCategory.objects.filter(name=category_name).first()
        if category_obj:
            course.category = category_obj

        thumbnail_file = request.FILES.get('thumbnail')
        if thumbnail_file:
            course.thumbnail.save(thumbnail_file.name, thumbnail_file)

        # ✅ Max lessons
        max_lessons_input = request.POST.get('max_lessons')
        lessons_count = course.lessons.count()  # Utilise le related_name="lessons"
        if not max_lessons_input or not max_lessons_input.isdigit() or int(max_lessons_input) < lessons_count:
            messages.error(
                request,
                f"Max lessons must be a number greater than or equal to current number of lessons ({lessons_count})."
            )
            return render(request, 'courses/course_edit.html', {
                'course': course,
                'categories': categories,
            })
        course.max_lessons = int(max_lessons_input)

        course.save()
        messages.success(request, "Course updated successfully!")
        return redirect('courses')

    return render(request, 'courses/course_edit.html', {
        'course': course,
        'categories': categories,
    })

@login_required
def delete_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    course.delete()
    messages.success(request, "Course deleted successfully.")
    return redirect('courses')  # Redirige vers la page liste des cours

@csrf_exempt
def schedule_course(request, course_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            date_str = data.get("publish_date")
            if not date_str:
                return JsonResponse({"error": "Date manquante"}, status=400)

            # Conversion string -> datetime
            publish_date = timezone.datetime.strptime(date_str, "%Y-%m-%d %H:%M")
            publish_date = timezone.make_aware(publish_date)

            course = Course.objects.get(id=course_id)
            course.publish_date = publish_date
            course.visible = False  # sera publié automatiquement plus tard
            course.save(update_fields=['publish_date', 'visible'])

            return JsonResponse({"success": True})
        except Course.DoesNotExist:
            return JsonResponse({"error": "Course not found"}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)


@login_required
def course_detail(request, course_id):
    # Vérifier que l'utilisateur est teacher
    if request.user.role != 'teacher':
        return redirect('unauthorized')

    course = get_object_or_404(Course, pk=course_id, teacher=request.user)

    # Leçons
    lessons = Lesson.objects.filter(course=course).order_by('order')

    # Préparer les ressources des leçons et sublessons
    import json
    from django.core.serializers.json import DjangoJSONEncoder

    for lesson in lessons:
        lesson.resource_list = Resource.objects.filter(lesson=lesson, sub_lesson__isnull=True)
        lesson.resources_json = json.dumps(
            list(lesson.resource_list.values('id', 'title', 'file', 'resource_type')),
            cls=DjangoJSONEncoder
        )

        sub_lessons = SubLesson.objects.filter(lesson=lesson).order_by('order')
        for sub in sub_lessons:
            sub.resource_list = Resource.objects.filter(sub_lesson=sub)
            sub.resources_json = json.dumps(
                list(sub.resource_list.values('id', 'title', 'file', 'resource_type')),
                cls=DjangoJSONEncoder
            )
        lesson.sublessons = sub_lessons

    # Étudiants
    students = course.students.all()
    all_students = User.objects.filter(role='student').exclude(id__in=students)

    # Durée
    if course.start_date and course.end_date:
        total_days = (course.end_date - course.start_date).days
        duration_weeks = total_days // 7
        duration_days = total_days % 7
    else:
        duration_weeks = duration_days = None

    # ✅ Reviews
    reviews = course.reviews.all().order_by('-created_at')  # les plus récents d'abord
    for review in reviews:
        review.full_stars = review.rating
        review.empty_stars = 5 - review.rating
        # Likes
        review.likes_count = review.likes.count()
        review.user_liked = request.user in review.likes.all()

        # Temps relatif type "2 hours ago"
        delta = timezone.now() - review.created_at
        if delta.days >= 1:
            review.time_ago = f"{delta.days} day{'s' if delta.days > 1 else ''} ago"
        elif delta.seconds >= 3600:
            hours = delta.seconds // 3600
            review.time_ago = f"{hours} hour{'s' if hours > 1 else ''} ago"
        elif delta.seconds >= 60:
            minutes = delta.seconds // 60
            review.time_ago = f"{minutes} minute{'s' if minutes > 1 else ''} ago"
        else:
            review.time_ago = "Just now"

    # ⭐ Moyenne des étoiles
    avg_rating = course.reviews.aggregate(avg=Avg('rating'))['avg'] or 0
    avg_full_stars = int(avg_rating)
    avg_empty_stars = 5 - avg_full_stars

    return render(request, 'courses/detailsCourse.html', {
        'course': course,
        'lessons': lessons,
        'students': students,
        'all_students': all_students,
        'duration_weeks': duration_weeks,
        'duration_days': duration_days,
        'reviews': reviews,
        'avg_rating': avg_rating,
        'avg_full_stars': avg_full_stars,
        'avg_empty_stars': avg_empty_stars,
    })



@login_required
def teacher_groupes(request):
    if request.user.role != 'teacher':
        return redirect('unauthorized')

    groupes = GroupeEtude.objects.filter(createur=request.user)
    return render(request, 'etude/groupes.html', {'groupes': groupes})

# --------------------------
# Add etude group 
# --------------------------
@login_required
def add_groupe(request):
    if request.user.role != 'teacher':
        return redirect('unauthorized')

    if request.method == 'POST':
        nom = request.POST.get('nom')
        description = request.POST.get('description')

        if nom:
            GroupeEtude.objects.create(
                nom=nom,
                description=description,
                createur=request.user
            )
            messages.success(request, "Study group created successfully!")
            return redirect('teacher_groups')  

    return render(request, 'etude/addGroupe.html')
    # --------------------------
# Edit a group
# --------------------------
@login_required
def edit_group(request, group_id):
    if request.user.role != 'teacher':
        return redirect('unauthorized')

    groupe = get_object_or_404(GroupeEtude, id=group_id, createur=request.user)

    if request.method == "POST":
        form = GroupeEtudeForm(request.POST, instance=groupe)
        if form.is_valid():
            form.save()
            messages.success(request, "Group updated successfully!")
            return redirect('teacher_groups')  # redirect to alias name used in templates
    else:
        form = GroupeEtudeForm(instance=groupe)

    # Render the actual template you have in TeacherDash/templates/etude/
    return render(request, 'etude/editGroupe.html', {'form': form, 'groupe': groupe})


@login_required
def delete_group(request, group_id):
    if request.user.role != 'teacher':
        return redirect('unauthorized')

    groupe = get_object_or_404(GroupeEtude, id=group_id, createur=request.user)

    if request.method == "POST":
        groupe.delete()
        messages.success(request, "Group deleted successfully!")
        return redirect('teacher_groups')

    return render(request, 'etude/deleteGroupe.html', {'groupe': groupe})


# etude resources
@login_required
def add_resource_etude(request, group_id):
    # only the group creator (teacher) can upload resources
    if request.user.role != 'teacher':
        return redirect('unauthorized')
    groupe = get_object_or_404(GroupeEtude, id=group_id, createur=request.user)
    if request.method == "POST":
        form = ResourceEtudeForm(request.POST, request.FILES)
        if form.is_valid():
            res = form.save(commit=False)
            res.groupe = groupe
            res.uploaded_by = request.user
            res.save()
            messages.success(request, "Resource uploaded.")
            return redirect('teacher_groups')
    else:
        form = ResourceEtudeForm()
    return render(request, 'etude/addResourceEtude.html', {'form': form, 'groupe': groupe})

@login_required
def teacher_group_detail(request, group_id):
    if request.user.role != 'teacher':
        return redirect('unauthorized')
    groupe = get_object_or_404(GroupeEtude, id=group_id, createur=request.user)
    members = groupe.membres.all()
    resources = groupe.resources_etude.all()
    messages_list = groupe.messages.all().order_by('date_envoi')
    return render(request, 'etude/groupDetail.html', {
        'groupe': groupe,
        'members': members,
        'resources': resources
    })

@login_required
def add_meeting(request, group_id):
    groupe = get_object_or_404(GroupeEtude, id=group_id, createur=request.user)
    if request.method == "POST":
        form = MeetingForm(request.POST)
        if form.is_valid():
            meeting = form.save(commit=False)
            meeting.groupe = groupe
            meeting.created_by = request.user
            # Generate unique Jitsi Meet link
            import uuid
            unique_id = str(uuid.uuid4())[:8]
            meeting.meet_link = f"https://meet.jit.si/{unique_id}"
            meeting.save()
            messages.success(request, f"Meeting created! Join link: {meeting.meet_link}")
            return redirect('teacher_group_detail', group_id=groupe.id)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = MeetingForm()
    return render(request, 'etude/addMeeting.html', {'form': form, 'groupe': groupe})

@login_required
def meetings_json(request, group_id):
    groupe = get_object_or_404(GroupeEtude, id=group_id)
    qs = groupe.meetings.all()
    events = []
    for m in qs:
        events.append({
            "id": m.id,
            "title": m.title,
            "start": m.start.isoformat(),
            "end": m.end.isoformat(),
            "url": m.meet_link,  
            "extendedProps": {
                "meet_link": m.meet_link
            }
        })
    return JsonResponse(events, safe=False)

@login_required
def group_meetings(request, group_id):
    groupe = get_object_or_404(GroupeEtude, id=group_id, createur=request.user)
    meetings = groupe.meetings.all().order_by('start')
    return render(request, 'etude/groupMeetings.html', {
        'groupe': groupe,
        'meetings': meetings
    })

@login_required
def meeting_detail(request, group_id, meeting_id):
    groupe = get_object_or_404(GroupeEtude, id=group_id, createur=request.user)
    meeting = get_object_or_404(Meeting, id=meeting_id, groupe=groupe)
    return render(request, 'etude/meeting_detail.html', {
        'groupe': groupe,
        'meeting': meeting
    })


@login_required
def group_meetings_by_date(request, group_id, date):
    """
    Show meetings for group filtered by date (expected format: YYYY-MM-DD).
    """
    groupe = get_object_or_404(GroupeEtude, id=group_id, createur=request.user)
    dt = parse_date(date)
    if not dt:
        messages.error(request, "Invalid date format. Use YYYY-MM-DD.")
        return redirect('group_meetings', group_id=group_id)

    meetings = groupe.meetings.filter(start__date=dt).order_by('start')
    return render(request, 'etude/groupMeetings.html', {'groupe': groupe, 'meetings': meetings})


@login_required
def group_meetings_by_time(request, group_id, time):
    """
    Show meetings for group filtered by time (expected format: HH:MM).
    This filters meetings whose start time matches the provided time.
    """
    groupe = get_object_or_404(GroupeEtude, id=group_id, createur=request.user)
    t = parse_time(time)
    if not t:
        messages.error(request, "Invalid time format. Use HH:MM.")
        return redirect('group_meetings', group_id=group_id)

    meetings = groupe.meetings.filter(start__time=t).order_by('start')
    return render(request, 'etude/groupMeetings.html', {'groupe': groupe, 'meetings': meetings})


@login_required
def group_meetings_by_date_time(request, group_id, date, time):
    """
    Filter meetings by both date and time. date=YYYY-MM-DD, time=HH:MM
    """
    groupe = get_object_or_404(GroupeEtude, id=group_id, createur=request.user)
    dt = parse_date(date)
    t = parse_time(time)
    if not dt or not t:
        messages.error(request, "Invalid date/time format. Use YYYY-MM-DD and HH:MM.")
        return redirect('group_meetings', group_id=group_id)

    # combine into a datetime for exact match on start
    start_dt = datetime.datetime.combine(dt, t)
    meetings = groupe.meetings.filter(start=start_dt).order_by('start')
    return render(request, 'etude/groupMeetings.html', {'groupe': groupe, 'meetings': meetings})
    


@login_required
def add_lesson(request):
    if request.method == 'POST':
        course_id = request.POST.get('course_id')
        course = get_object_or_404(Course, id=course_id, teacher=request.user)
        title = request.POST.get('title')
        description = request.POST.get('description', '')
        max_sublessons = request.POST.get('max_sublessons')

        lesson = Lesson.objects.create(
            course=course,
            title=title,
            description=description,
            order=Lesson.objects.filter(course=course).count() + 1,
            max_sublessons=int(max_sublessons) if max_sublessons and max_sublessons.isdigit() else None
        )

        # fichiers uploadés
        for file in request.FILES.getlist('resources'):
            ext = file.name.split('.')[-1].lower()
            if ext in ['jpg', 'jpeg', 'png', 'gif', 'bmp']:
                r_type = 'image'
            elif ext == 'pdf':
                r_type = 'pdf'
            elif ext in ['mp4', 'mov', 'avi']:
                r_type = 'video'
            elif ext in ['mp3', 'wav']:
                r_type = 'audio'
            else:
                r_type = 'other'

            Resource.objects.create(
                lesson=lesson,
                title=file.name,
                resource_type=r_type,
                file=file
            )

        # URLs externes
        external_urls = request.POST.getlist('external_url[]')
        for url in external_urls:
            if url.strip():
                Resource.objects.create(
                    lesson=lesson,
                    title=url.split('/')[-1],
                    external_url=url,
                    resource_type='external'
                )

        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def add_sublesson(request):
    if request.method == 'POST':
        lesson_id = request.POST.get('lesson_id')
        lesson = get_object_or_404(Lesson, id=lesson_id)
        title = request.POST.get('title')
        content = request.POST.get('content', '')

        sublesson = SubLesson.objects.create(
            lesson=lesson,
            title=title,
            content=content,
            order=SubLesson.objects.filter(lesson=lesson).count() + 1
        )

        # fichiers uploadés
        for file in request.FILES.getlist('resources'):
            ext = file.name.split('.')[-1].lower()
            if ext in ['jpg', 'jpeg', 'png', 'gif', 'bmp']:
                r_type = 'image'
            elif ext == 'pdf':
                r_type = 'pdf'
            elif ext in ['mp4', 'mov', 'avi']:
                r_type = 'video'
            elif ext in ['mp3', 'wav']:
                r_type = 'audio'
            else:
                r_type = 'other'

            Resource.objects.create(
                sub_lesson=sublesson,
                title=file.name,
                resource_type=r_type,
                file=file
            )

        # ✅ URLs externes
        external_urls = request.POST.getlist('external_url[]')
        for url in external_urls:
            if url.strip():
                Resource.objects.create(
                    sub_lesson=sublesson,
                    title=url.split('/')[-1],
                    external_url=url,
                    resource_type='external'
                )

        return JsonResponse({'success': True})
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

# Delete Lesson
@login_required
def delete_lesson(request, lesson_id):
    if request.method == 'POST':
        lesson = get_object_or_404(Lesson, id=lesson_id, course__teacher=request.user)
        lesson.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)

# Delete SubLesson
@login_required
def delete_sublesson(request, sublesson_id):
    if request.method == 'POST':
        sub = get_object_or_404(SubLesson, id=sublesson_id, lesson__course__teacher=request.user)
        sub.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)

# Update Lesson
@login_required
def update_lesson(request):
    if request.method == 'POST':
        lesson_id = request.POST.get('lesson_id')
        lesson = get_object_or_404(Lesson, id=lesson_id, course__teacher=request.user)

        # Mise à jour basique
        lesson.title = request.POST.get('title')
        lesson.description = request.POST.get('description')
        
        max_sublessons = request.POST.get('max_sublessons')
        lesson.max_sublessons = int(max_sublessons) if max_sublessons and max_sublessons.isdigit() else None

        lesson.save()

        # Suppression des ressources cochées
        resources_to_delete = request.POST.getlist('delete_resources[]')
        if resources_to_delete:
            Resource.objects.filter(id__in=resources_to_delete).delete()

        # Ajout des URLs externes
        external_urls = request.POST.getlist('external_url[]')
        for url in external_urls:
            if url.strip():
                Resource.objects.create(
                    lesson=lesson,
                    title=url.split('/')[-1],
                    external_url=url,
                    resource_type='external'
                )

        # Ajout des nouvelles ressources uploadées
        for file in request.FILES.getlist('resources'):
            ext = file.name.split('.')[-1].lower()
            if ext in ['jpg', 'jpeg', 'png', 'gif', 'bmp']:
                r_type = 'image'
            elif ext == 'pdf':
                r_type = 'pdf'
            elif ext in ['mp4', 'mov', 'avi']:
                r_type = 'video'
            elif ext in ['mp3', 'wav']:
                r_type = 'audio'
            else:
                r_type = 'other'

            Resource.objects.create(
                lesson=lesson,
                title=file.name,
                resource_type=r_type,
                file=file
            )

        return JsonResponse({
            'success': True,
            'lesson': {
                'id': lesson.id,
                'title': lesson.title,
                'description': lesson.description,
                'max_sublessons': lesson.max_sublessons
            }
        })
 
    return JsonResponse({'success': False}, status=400)


@login_required
def update_sublesson(request):
    if request.method == 'POST':
        sublesson_id = request.POST.get('sublesson_id')
        sub = get_object_or_404(SubLesson, id=sublesson_id, lesson__course__teacher=request.user)

        # Mise à jour des champs de base
        sub.title = request.POST.get('title')
        sub.content = request.POST.get('content')
        sub.save()

        # ✅ Suppression des ressources cochées ou marquées pour suppression
        resources_to_delete = request.POST.getlist('delete_resources[]')
        if resources_to_delete:
            Resource.objects.filter(id__in=resources_to_delete, sub_lesson=sub).delete()

        # ✅ Ajout des URLs externes
        external_urls = request.POST.getlist('external_url[]')
        for url in external_urls:
            if url.strip():
                Resource.objects.create(
                    sub_lesson=sub,
                    title=url.split('/')[-1],
                    external_url=url,
                    resource_type='external'
                )

        # ✅ Ajout des nouvelles ressources uploadées
        for file in request.FILES.getlist('resources'):
            ext = file.name.split('.')[-1].lower()
            if ext in ['jpg', 'jpeg', 'png', 'gif', 'bmp']:
                r_type = 'image'
            elif ext == 'pdf':
                r_type = 'pdf'
            elif ext in ['mp4', 'mov', 'avi']:
                r_type = 'video'
            elif ext in ['mp3', 'wav']:
                r_type = 'audio'
            else:
                r_type = 'other'

            Resource.objects.create(
                sub_lesson=sub,
                title=file.name,
                resource_type=r_type,
                file=file
            )

        return JsonResponse({'success': True})

    return JsonResponse({'success': False}, status=400)


def get_lesson_resources(request, lesson_id):
    lesson = get_object_or_404(Lesson, id=lesson_id, course__teacher=request.user)
    resources = list(lesson.resources.values('id', 'title', 'file'))
    return JsonResponse({'resources': resources})

@login_required
@csrf_exempt
def toggle_visibility(request, type, id):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Méthode non autorisée"})

    try:
        data = json.loads(request.body)
        visible = data.get("visible", False)

        from core.models import Course, Lesson, SubLesson
        model_map = {
            "course": Course,
            "lesson": Lesson,
            "sublesson": SubLesson
        }

        model = model_map.get(type)
        if not model:
            return JsonResponse({"success": False, "error": "Type invalide"})

        obj = model.objects.get(id=id)
        obj.visible = visible
        obj.save()

        return JsonResponse({"success": True, "visible": obj.visible})
    except Exception as e:
        print("❌ Erreur toggle_visibility:", e)
        return JsonResponse({"success": False, "error": str(e)})

@login_required
def assign_students_to_course(request, course_id):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        student_ids = data.get('students', [])  # ⚠️ doit matcher le body JS
        course = get_object_or_404(Course, id=course_id)
        course.students.add(*User.objects.filter(id__in=student_ids, role='student'))
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)

@login_required
def remove_student_from_course(request, course_id):
    if request.method == 'POST':
        import json
        data = json.loads(request.body)
        student_id = data.get('student_id')
        course = get_object_or_404(Course, id=course_id)

        student = get_object_or_404(User, id=student_id, role='student')
        course.students.remove(student)

        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=400)


@login_required
def toggle_like_review(request, review_id):
    review = get_object_or_404(Review, id=review_id)
    user = request.user

    if user in review.likes.all():
        review.likes.remove(user)
        liked = False
    else:
        review.likes.add(user)
        liked = True

    return JsonResponse({'liked': liked, 'likes_count': review.likes.count()})


@login_required
def delete_review(request, review_id):
    review = get_object_or_404(Review, id=review_id)

    # Seul le propriétaire ou un professeur peut supprimer
    if request.user == review.student or request.user.role == 'teacher':
        review.delete()
        return JsonResponse({'deleted': True})
    else:
        return JsonResponse({'deleted': False, 'error': 'Unauthorized'}, status=403)


@method_decorator(login_required, name='dispatch')
class SpecialityListView(View):
    def get(self, request):
        specialities = Speciality.objects.all()
        return render(request, 'specialities/listSpecialities.html', {'specialities': specialities})

@method_decorator(login_required, name='dispatch')
class SpecialityCreateView(View):
    def get(self, request):
        form = SpecialityForm()
        return render(request, 'specialities/addSpecialities.html', {'form': form})

    def post(self, request):
        form = SpecialityForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('specialities')  # corriger nom url
        return render(request, 'specialities/addSpecialities.html', {'form': form})

@method_decorator(login_required, name='dispatch')
class SpecialityUpdateView(View):
    def get(self, request, speciality_id):
        speciality = get_object_or_404(Speciality, pk=speciality_id)
        form = SpecialityForm(instance=speciality)
        return render(request, 'specialities/editSpecialities.html', {'form': form})

    def post(self, request, speciality_id):
        speciality = get_object_or_404(Speciality, pk=speciality_id)
        form = SpecialityForm(request.POST, instance=speciality)
        if form.is_valid():
            form.save()
            return redirect('specialities')
        return render(request, 'specialities/editSpecialities.html', {'form': form})

@method_decorator(login_required, name='dispatch')
class SpecialityDeleteView(View):
    def post(self, request, speciality_id):
        speciality = get_object_or_404(Speciality, pk=speciality_id)
        speciality.delete()
        return redirect('specialities')

class ListCertificatView(View):
    def get(self, request):
        search_query = request.GET.get('search', '')
        certificates_qs = Certificate.objects.annotate(
            exercise_count=Count('exercises'),
            participant_count=Count('certificateattempt', distinct=True),
            succeeded_count=Count('certificateattempt', filter=Q(certificateattempt__passed=True), distinct=True)
        ).all().order_by('id')

        if search_query:
            certificates_qs = certificates_qs.filter(
                Q(title__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(speciality__name__icontains=search_query)
            )

        if search_query:
            certificates = certificates_qs
            page_obj = None
        else:
            paginator = Paginator(certificates_qs, 10)  # 10 items per page
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)
            certificates = page_obj.object_list

        return render(request, 'certificats/listCertificate.html', {
            'certificates': certificates,
            'page_obj': page_obj,
            'search_query': search_query
        })


@login_required
def create_certificate(request):
    CertificatExerciseFormSet = modelformset_factory(
        CertificatExercise,
        form=CertificatExerciseForm,
        extra=1,  # une ligne vide par défaut
        can_delete=True
    )

    exercise_error = None

    if request.method == 'POST':
        cert_form = CertificateForm(request.POST, request.FILES)
        formset = CertificatExerciseFormSet(
            request.POST,
            queryset=CertificatExercise.objects.none(),
            prefix='exercise'
        )

        if cert_form.is_valid() and formset.is_valid():
            exercises_to_save = []
            has_empty_exercise = False

            for form in formset:
                if not form.cleaned_data.get('DELETE'):
                    exercise_type = form.cleaned_data.get('exercise_type')
                    has_data = (
                        form.cleaned_data.get('question') or
                        form.cleaned_data.get('option1') or
                        form.cleaned_data.get('option2') or
                        form.cleaned_data.get('option3') or
                        form.cleaned_data.get('option4')
                    )
                    if has_data and not exercise_type:
                        has_empty_exercise = True
                        break
                    elif exercise_type:
                        exercises_to_save.append(form)

            if has_empty_exercise:
                exercise_error = "All exercises must have an exercise type selected."
            elif not exercises_to_save:
                exercise_error = "At least one exercise is required."
            else:
                certificate = cert_form.save()
                for form in exercises_to_save:
                    exercise = form.save(commit=False)
                    exercise.certificate = certificate
                    exercise.save()
                messages.success(request, "Certificate created successfully!")
                return redirect('list_certificates')
        else:
            exercise_error = "Please correct the errors in the exercises."
    else:
        cert_form = CertificateForm()
        formset = CertificatExerciseFormSet(
            queryset=CertificatExercise.objects.none(),
            prefix='exercise'
        )

    return render(request, 'certificats/certificate_form.html', {
        'cert_form': cert_form,
        'formset': formset,
        'exercise_error': exercise_error,
    })


# AJOUT / MODIF: edit_certificate (passer et traiter formset avec exercises existants)
@login_required
def edit_certificate(request, cert_id):
    certificate = get_object_or_404(Certificate, pk=cert_id)
    CertificatExerciseFormSet = modelformset_factory(
        CertificatExercise,
        form=CertificatExerciseForm,
        extra=0,  # pas de formulaire vide automatiquement
        can_delete=True
    )

    exercise_error = None

    if request.method == "POST":
        cert_form = CertificateForm(request.POST, request.FILES, instance=certificate)
        formset = CertificatExerciseFormSet(
            request.POST,
            queryset=CertificatExercise.objects.filter(certificate=certificate),
            prefix='exercise'
        )

        if cert_form.is_valid() and formset.is_valid():
            exercises_to_save = []
            has_empty_exercise = False

            for form in formset:
                if not form.cleaned_data.get('DELETE'):
                    exercise_type = form.cleaned_data.get('exercise_type')
                    has_data = (
                        form.cleaned_data.get('question') or
                        form.cleaned_data.get('option1') or
                        form.cleaned_data.get('option2') or
                        form.cleaned_data.get('option3') or
                        form.cleaned_data.get('option4')
                    )
                    if has_data and not exercise_type:
                        has_empty_exercise = True
                        break
                    elif exercise_type:
                        exercises_to_save.append(form)

            if has_empty_exercise:
                exercise_error = "All exercises must have an exercise type selected."
            elif not exercises_to_save:
                exercise_error = "At least one exercise is required."
            else:
                try:
                    with transaction.atomic():
                        cert_form.save()
                        for form in formset.forms:
                            if form.cleaned_data.get('DELETE') and form.instance.pk:
                                form.instance.delete()
                            elif form.cleaned_data.get('exercise_type'):
                                exercise = form.save(commit=False)
                                exercise.certificate = certificate
                                exercise.save()
                        messages.success(request, "Certificate updated successfully!")
                        return redirect('list_certificates')
                except Exception as e:
                    messages.error(request, f"Error updating certificate: {str(e)}")
        else:
            exercise_error = "Please correct the errors in the exercises."
    else:
        cert_form = CertificateForm(instance=certificate)
        formset = CertificatExerciseFormSet(
            queryset=CertificatExercise.objects.filter(certificate=certificate),
            prefix='exercise'
        )

    return render(request, 'certificats/editCertificate.html', {
        'cert_form': cert_form,
        'formset': formset,
        'certificate': certificate,
        'exercise_error': exercise_error,
    })

# AJOUT: delete_certificate
@login_required
def delete_certificate(request, cert_id):
    if request.user.role not in ['admin', 'teacher']:
        return redirect(reverse('unauthorized'))
    certificate = get_object_or_404(Certificate, pk=cert_id)
    if request.method == "POST":
        certificate.delete()
        messages.success(request, "Certificate deleted successfully.")
        return redirect('list_certificates')
    # If not POST, redirect back to list
    return redirect('list_certificates')

@login_required
def preview_certificate(request, cert_id):
    if request.user.role not in ['admin', 'teacher']:
        return redirect(reverse('unauthorized'))
    certificate = get_object_or_404(Certificate, pk=cert_id)
    exercises = CertificatExercise.objects.filter(certificate=certificate)
    return render(request, 'certificats/preview_certificate.html', {
        'certificate': certificate,
        'exercises': exercises,
    })

@login_required
def certificate_results(request, cert_id):
    if request.user.role not in ['admin', 'teacher']:
        return redirect(reverse('unauthorized'))
    certificate = get_object_or_404(Certificate, pk=cert_id)
    attempts = CertificateAttempt.objects.filter(certificate=certificate).select_related('user').order_by('-completed_at')
    return render(request, 'certificats/certificate_results.html', {
        'certificate': certificate,
        'attempts': attempts,
    })

@login_required
def attempt_details(request, attempt_id):
    if request.user.role not in ['admin', 'teacher']:
        return redirect(reverse('unauthorized'))
    attempt = get_object_or_404(CertificateAttempt, pk=attempt_id)
    answers = attempt.answers.select_related('exercise').all()
    return render(request, 'certificats/attempt_details.html', {
        'attempt': attempt,
        'answers': answers,
    })

@login_required
def export_certificate_results_csv(request, cert_id):
    if request.user.role not in ['admin', 'teacher']:
        return redirect(reverse('unauthorized'))
    certificate = get_object_or_404(Certificate, pk=cert_id)
    attempts = CertificateAttempt.objects.filter(certificate=certificate).select_related('user').prefetch_related('answers__exercise').order_by('-completed_at')

    # Get all exercises for this certificate
    exercises = list(CertificatExercise.objects.filter(certificate=certificate).order_by('id'))

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Certificate Results"

    # Write headers
    headers = ['User', 'Score', 'Total Questions', 'Passed', 'Completed At']
    for exercise in exercises:
        headers.append(f"{exercise.question} - Answer")
        headers.append(f"{exercise.question} - Correct Answer")

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Write data
    for row_num, attempt in enumerate(attempts, 2):
        ws.cell(row=row_num, column=1, value=attempt.user.username)
        ws.cell(row=row_num, column=2, value=attempt.score)
        ws.cell(row=row_num, column=3, value=attempt.total_questions)
        ws.cell(row=row_num, column=4, value='Yes' if attempt.passed else 'No')
        ws.cell(row=row_num, column=5, value=attempt.completed_at.strftime('%Y-%m-%d %H:%M:%S'))

        # Create a dict of exercise_id to answer for quick lookup
        answers_dict = {answer.exercise_id: answer for answer in attempt.answers.all()}

        col_num = 6
        for exercise in exercises:
            if exercise.id in answers_dict:
                answer = answers_dict[exercise.id]
                ws.cell(row=row_num, column=col_num, value=answer.answer)
                ws.cell(row=row_num, column=col_num + 1, value=exercise.correct_answer)
            else:
                ws.cell(row=row_num, column=col_num, value='')
                ws.cell(row=row_num, column=col_num + 1, value=exercise.correct_answer)
            col_num += 2

    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{certificate.title}_results.xlsx"'

    wb.save(response)
    return response




@login_required
@csrf_exempt  # facultatif si tu passes bien le token CSRF dans fetch
def delete_resource(request, resource_id):
    # Vérifie que la requête simule bien une suppression
    if request.method == 'POST' and request.POST.get('_method') == 'DELETE':
        resource = get_object_or_404(Resource, id=resource_id, sub_lesson__lesson__course__teacher=request.user)
        resource.delete()
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=400)

